from flask import render_template, request, redirect, url_for, flash, session, send_file
from flask_login import login_user, logout_user, login_required, current_user
from datetime import datetime, timedelta
from app import app
from models import User, Book, Transaction, ActivityLog, BookRequest, PaymentRecord, users, books, transactions, activity_logs, book_requests, payment_records, system_settings

@app.context_processor
def inject_system_settings():
    """Make system settings available to all templates"""
    from datetime import datetime
    return dict(system_settings=system_settings, current_year=datetime.now().year)

def log_activity(user_id, action, ip_address, details=None):
    """Log user activity"""
    activity = ActivityLog(user_id, action, ip_address, details)
    activity_logs.append(activity)

def get_client_ip():
    """Get client IP address"""
    if request.environ.get('HTTP_X_FORWARDED_FOR') is None:
        return request.environ['REMOTE_ADDR']
    else:
        return request.environ['HTTP_X_FORWARDED_FOR']

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.get_by_username(username)
        if user and user.role == 'admin' and user.check_password(password):
            login_user(user)
            log_activity(user.id, 'admin_login', get_client_ip())
            flash('Admin login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid admin credentials!', 'danger')

    return render_template('admin_login.html')

@app.route('/teacher/login', methods=['GET', 'POST'])
def teacher_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.get_by_username(username)
        if user and user.role == 'teacher' and user.check_password(password):
            login_user(user)
            log_activity(user.id, 'teacher_login', get_client_ip())
            flash('Teacher login successful!', 'success')
            return redirect(url_for('teacher_dashboard'))
        else:
            flash('Invalid teacher credentials!', 'danger')

    return render_template('teacher_login.html')

@app.route('/student/login', methods=['GET', 'POST'])
def student_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = User.get_by_username(username)
        if user and user.role == 'student' and user.check_password(password):
            login_user(user)
            log_activity(user.id, 'student_login', get_client_ip())
            flash('Student login successful!', 'success')
            return redirect(url_for('student_dashboard'))
        else:
            flash('Invalid student credentials!', 'danger')

    return render_template('student_login.html')

@app.route('/logout')
@login_required
def logout():
    log_activity(current_user.id, 'logout', get_client_ip())
    logout_user()
    flash('You have been logged out successfully!', 'info')
    return redirect(url_for('index'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Dashboard statistics
    total_books = len(books)
    total_users = len([u for u in users.values() if u.role != 'admin'])
    active_transactions = len([t for t in transactions.values() if not t.is_returned])
    overdue_books = sum(1 for t in transactions.values() if not t.is_returned and t.due_date < datetime.now())

    return render_template('admin_dashboard.html', 
                         total_books=total_books,
                         total_users=total_users,
                         active_transactions=active_transactions,
                         overdue_books=overdue_books)

@app.route('/teacher/dashboard')
@login_required
def teacher_dashboard():
    if current_user.role != 'teacher':
        flash('Access denied! Teacher privileges required.', 'danger')
        return redirect(url_for('index'))

    return render_template('teacher_dashboard.html')

@app.route('/student/dashboard')
@login_required
def student_dashboard():
    if current_user.role != 'student':
        flash('Access denied! Student privileges required.', 'danger')
        return redirect(url_for('index'))

    # Get student's borrowed books
    user_transactions = [t for t in transactions.values() if t.user_id == current_user.id and not t.is_returned]
    borrowed_books = []
    for transaction in user_transactions:
        book = books.get(transaction.book_id)
        if book:
            borrowed_books.append({
                'book': book,
                'transaction': transaction,
                'is_overdue': transaction.due_date < datetime.now()
            })

    return render_template('student_dashboard.html', borrowed_books=borrowed_books)

@app.route('/admin/books')
@login_required
def book_management():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    return render_template('book_management.html', books=books.values())

@app.route('/admin/books/add', methods=['POST'])
@login_required
def add_book():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    title = request.form['title']
    author = request.form['author']
    isbn = request.form['isbn']
    genre = request.form['genre']
    publication_year = int(request.form['publication_year'])
    total_copies = int(request.form['total_copies'])

    book_id = f"book{len(books) + 1}"
    book = Book(book_id, title, author, isbn, genre, publication_year, total_copies)
    books[book_id] = book

    log_activity(current_user.id, 'add_book', get_client_ip(), {'book_title': title})
    flash(f'Book "{title}" added successfully!', 'success')
    return redirect(url_for('book_management'))

@app.route('/admin/books/delete/<book_id>')
@login_required
def delete_book(book_id):
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    if book_id in books:
        book_title = books[book_id].title
        del books[book_id]
        log_activity(current_user.id, 'delete_book', get_client_ip(), {'book_title': book_title})
        flash(f'Book "{book_title}" deleted successfully!', 'success')
    else:
        flash('Book not found!', 'danger')

    return redirect(url_for('book_management'))

@app.route('/admin/users')
@login_required
def user_management():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    non_admin_users = [u for u in users.values() if u.role != 'admin']
    return render_template('user_management.html', users=non_admin_users)

@app.route('/admin/users/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    username = request.form['username']
    email = request.form['email']
    password = request.form['password']
    role = request.form['role']
    full_name = request.form['full_name']
    admission_number = request.form.get('admission_number', '') if role == 'student' else None

    # For students, use admission number as username if provided
    if role == 'student' and admission_number:
        username = admission_number

    # Check if username already exists
    if User.get_by_username(username):
        flash('Username already exists!', 'danger')
        return redirect(url_for('user_management'))

    user_id = f"{role}{len([u for u in users.values() if u.role == role]) + 1}"
    user = User(user_id, username, email, password, role, full_name, admission_number)
    users[user_id] = user

    log_activity(current_user.id, 'add_user', get_client_ip(), {'username': username, 'role': role})
    flash(f'User "{username}" added successfully!', 'success')
    return redirect(url_for('user_management'))

@app.route('/admin/activity')
@login_required
def activity_monitoring():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Get recent activity logs (last 50)
    recent_logs = sorted(activity_logs, key=lambda x: x.timestamp, reverse=True)[:50]

    # Enrich logs with user information
    enriched_logs = []
    for log in recent_logs:
        user = users.get(log.user_id)
        enriched_logs.append({
            'log': log,
            'user': user
        })

    return render_template('activity_monitoring.html', logs=enriched_logs)

@app.route('/books/search')
@login_required
def book_search():
    query = request.args.get('q', '')
    genre = request.args.get('genre', '')

    filtered_books = []
    for book in books.values():
        if query.lower() in book.title.lower() or query.lower() in book.author.lower():
            if not genre or book.genre.lower() == genre.lower():
                filtered_books.append(book)

    genres = list(set(book.genre for book in books.values()))
    return render_template('book_search.html', books=filtered_books, query=query, genres=genres, selected_genre=genre)

@app.route('/books/request/<book_id>')
@login_required
def request_book(book_id):
    if current_user.role not in ['student', 'teacher']:
        flash('Only students and teachers can request books!', 'danger')
        return redirect(url_for('book_search'))

    book = books.get(book_id)
    if not book:
        flash('Book not found!', 'danger')
        return redirect(url_for('book_search'))

    # Check if user already has an active borrowing or pending request for this book
    active_transaction = [t for t in transactions.values() if t.user_id == current_user.id and t.book_id == book_id and not t.is_returned]
    pending_request = [r for r in book_requests.values() if r.user_id == current_user.id and r.book_id == book_id and r.status == 'pending']

    if active_transaction:
        flash('You have already borrowed this book!', 'warning')
        return redirect(url_for('book_search'))

    if pending_request:
        flash('You already have a pending request for this book!', 'warning')
        return redirect(url_for('book_search'))

    # Show request form
    return render_template('book_request_form.html', book=book)

@app.route('/books/request-quick/<book_id>')
@login_required
def request_book_quick(book_id):
    if current_user.role not in ['student', 'teacher']:
        flash('Only students and teachers can request books!', 'danger')
        return redirect(url_for('book_search'))

    book = books.get(book_id)
    if not book:
        flash('Book not found!', 'danger')
        return redirect(url_for('book_search'))

    # Check if user already has an active borrowing or pending request for this book
    active_transaction = [t for t in transactions.values() if t.user_id == current_user.id and t.book_id == book_id and not t.is_returned]
    pending_request = [r for r in book_requests.values() if r.user_id == current_user.id and r.book_id == book_id and r.status == 'pending']

    if active_transaction:
        flash('You have already borrowed this book!', 'warning')
        return redirect(url_for('book_search'))

    if pending_request:
        flash('You already have a pending request for this book!', 'warning')
        return redirect(url_for('book_search'))

    # Create book request
    request_id = f"req{len(book_requests) + 1}"
    book_request = BookRequest(request_id, current_user.id, book_id)
    book_requests[request_id] = book_request

    log_activity(current_user.id, 'request_book', get_client_ip(), {'book_title': book.title})
    flash(f'Request for "{book.title}" submitted successfully! Admin will review your request.', 'success')

    return redirect(url_for('student_dashboard' if current_user.role == 'student' else 'teacher_dashboard'))

@app.route('/books/request/<book_id>', methods=['POST'])
@login_required  
def request_book_with_message(book_id):
    if current_user.role not in ['student', 'teacher']:
        flash('Only students and teachers can request books!', 'danger')
        return redirect(url_for('book_search'))

    message = request.form.get('message', '')
    book = books.get(book_id)
    if not book:
        flash('Book not found!', 'danger')
        return redirect(url_for('book_search'))

    # Check for existing requests/borrowings
    active_transaction = [t for t in transactions.values() if t.user_id == current_user.id and t.book_id == book_id and not t.is_returned]
    pending_request = [r for r in book_requests.values() if r.user_id == current_user.id and r.book_id == book_id and r.status == 'pending']

    if active_transaction or pending_request:
        flash('You already have an active request or borrowing for this book!', 'warning')
        return redirect(url_for('book_search'))

    # Create book request with message
    request_id = f"req{len(book_requests) + 1}"
    book_request = BookRequest(request_id, current_user.id, book_id, message)
    book_requests[request_id] = book_request

    log_activity(current_user.id, 'request_book', get_client_ip(), {'book_title': book.title, 'message': message})
    flash(f'Request for "{book.title}" submitted successfully!', 'success')

    return redirect(url_for('student_dashboard' if current_user.role == 'student' else 'teacher_dashboard'))

@app.route('/admin/collect-fine/<transaction_id>', methods=['GET', 'POST'])
@login_required
def collect_fine(transaction_id):
    if current_user.role != 'admin':
        flash('Access denied! Only admin can collect fines.', 'danger')
        return redirect(url_for('index'))

    user_transaction = transactions.get(transaction_id)
    if not user_transaction:
        flash('Transaction not found!', 'danger')
        return redirect(url_for('borrow_return'))

    if user_transaction.fine_paid:
        flash('Fine has already been collected!', 'warning')
        return redirect(url_for('borrow_return'))

    fine_amount = user_transaction.update_fine()
    if fine_amount <= 0:
        flash('No fine due for this transaction!', 'info')
        return redirect(url_for('borrow_return'))

    if request.method == 'POST':
        payment_method = request.form.get('payment_method', 'cash')
        notes = request.form.get('notes', '')

        # Mark fine as paid
        user_transaction.fine_paid = True
        user_transaction.payment_date = datetime.now()
        user_transaction.payment_collected_by = current_user.id

        # Create payment record
        payment_id = f"pay{len(payment_records) + 1}"
        payment = PaymentRecord(payment_id, transaction_id, user_transaction.user_id, fine_amount)
        payment.collected_by = current_user.id
        payment.payment_method = payment_method
        payment.notes = notes
        payment_records[payment_id] = payment

        # Get user and book details
        user = users.get(user_transaction.user_id)
        book = books.get(user_transaction.book_id)

        log_activity(current_user.id, 'collect_fine', get_client_ip(), {
            'user': user.full_name if user else 'Unknown',
            'book': book.title if book else 'Unknown',
            'amount': f"{fine_amount} rs",
            'payment_method': payment_method
        })

        flash(f'Fine of {fine_amount} rs collected successfully from {user.full_name if user else "user"}!', 'success')
        return redirect(url_for('borrow_return'))

    # GET request - show collection form
    user = users.get(user_transaction.user_id)
    book = books.get(user_transaction.book_id)

    return render_template('collect_fine.html', 
                         transaction=user_transaction,
                         user=user,
                         book=book,
                         fine_amount=fine_amount)

@app.route('/admin/return-book/<transaction_id>')
@login_required
def admin_return_book(transaction_id):
    if current_user.role != 'admin':
        flash('Access denied! Only admin can return books.', 'danger')
        return redirect(url_for('index'))

    # Find the transaction
    user_transaction = transactions.get(transaction_id)

    if not user_transaction:
        flash('Transaction not found!', 'danger')
        return redirect(url_for('borrow_return'))

    if user_transaction.is_returned:
        flash('Book has already been returned!', 'warning')
        return redirect(url_for('borrow_return'))

    # Calculate fine if overdue
    fine_amount = user_transaction.update_fine()

    # Check if fine needs to be collected before return
    if fine_amount > 0 and not user_transaction.fine_paid:
        flash(f'Please collect fine of {fine_amount} rs before returning the book!', 'warning')
        return redirect(url_for('collect_fine', transaction_id=transaction_id))

    # Mark as returned
    user_transaction.is_returned = True
    user_transaction.return_date = datetime.now()

    # Update book availability
    book = books.get(user_transaction.book_id)
    if book:
        book.available_copies += 1

    # Get user details for logging
    user = users.get(user_transaction.user_id)

    # Log activity
    log_details = {
        'book_title': book.title if book else 'Unknown',
        'returned_for': user.full_name if user else 'Unknown User'
    }
    if fine_amount > 0:
        log_details['fine_collected'] = f"{fine_amount} rs"

    log_activity(current_user.id, 'admin_return_book', get_client_ip(), log_details)

    flash(f'Book returned successfully for {user.full_name if user else "user"}!', 'success')
    return redirect(url_for('borrow_return'))

@app.route('/admin/transactions')
@login_required
def borrow_return():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Get all transactions with book and user details
    enriched_transactions = []
    for transaction in transactions.values():
        user = users.get(transaction.user_id)
        book = books.get(transaction.book_id)

        enriched_transactions.append({
            'transaction': transaction,
            'user': user,
            'book': book,
            'is_overdue': not transaction.is_returned and transaction.due_date < datetime.now()
        })

    # Sort by transaction date (newest first)
    enriched_transactions.sort(key=lambda x: x['transaction'].transaction_date, reverse=True)

    # Update fines for all overdue transactions
    for item in enriched_transactions:
        item['transaction'].update_fine()

    return render_template('borrow_return.html', transactions=enriched_transactions)

@app.route('/admin/assign-book')
@login_required
def assign_book():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Get students and teachers
    eligible_users = [u for u in users.values() if u.role in ['student', 'teacher']]
    available_books = [b for b in books.values() if b.available_copies > 0]

    return render_template('assign_book.html', 
                         users=eligible_users,
                         books=available_books)

@app.route('/admin/assign-book', methods=['POST'])
@login_required
def assign_book_post():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    user_id = request.form['user_id']
    book_id = request.form['book_id']

    user = users.get(user_id)
    book = books.get(book_id)

    if not user or not book:
        flash('Invalid user or book selection!', 'danger')
        return redirect(url_for('assign_book'))

    if book.available_copies <= 0:
        flash('Book is not available for borrowing!', 'warning')
        return redirect(url_for('assign_book'))

    # Check if user already has this book
    user_transactions = [t for t in transactions.values() if t.user_id == user_id and t.book_id == book_id and not t.is_returned]
    if user_transactions:
        flash('User has already borrowed this book!', 'warning')
        return redirect(url_for('assign_book'))

    # Create transaction
    transaction_id = f"trans{len(transactions) + 1}"
    transaction = Transaction(transaction_id, user_id, book_id, 'borrow')
    transactions[transaction_id] = transaction

    # Update book availability
    book.available_copies -= 1

    log_activity(current_user.id, 'assign_book', get_client_ip(), {
        'book_title': book.title,
        'assigned_to': user.full_name,
        'user_role': user.role
    })

    flash(f'Book "{book.title}" assigned to {user.full_name} successfully! Due date: {transaction.due_date.strftime("%Y-%m-%d")}', 'success')
    return redirect(url_for('assign_book'))

@app.route('/admin/reports')
@login_required
def admin_reports():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Calculate comprehensive statistics
    total_books = len(books)
    total_copies = sum(book.total_copies for book in books.values())
    available_copies = sum(book.available_copies for book in books.values())
    borrowed_copies = total_copies - available_copies

    total_users = len([u for u in users.values() if u.role in ['student', 'teacher']])
    active_borrowers = len(set(t.user_id for t in transactions.values() if not t.is_returned))

    # Overdue books analysis
    overdue_transactions = []
    total_fines_due = 0.0
    total_fines_collected = 0.0

    for transaction in transactions.values():
        transaction.update_fine()
        if transaction.is_overdue and not transaction.is_returned:
            user = users.get(transaction.user_id)
            book = books.get(transaction.book_id)
            overdue_transactions.append({
                'transaction': transaction,
                'user': user,
                'book': book
            })
            if transaction.fine_paid:
                total_fines_collected += transaction.fine_amount
            else:
                total_fines_due += transaction.fine_amount

    # Calculate total revenue from all paid fines
    total_revenue = sum(payment.amount for payment in payment_records.values())

    # Calculate today's collections
    from datetime import date
    today = date.today()
    todays_collections = sum(
        payment.amount for payment in payment_records.values()
        if payment.payment_date.date() == today
    )

    # Most borrowed books
    book_borrow_count = {}
    for transaction in transactions.values():
        book_id = transaction.book_id
        book_borrow_count[book_id] = book_borrow_count.get(book_id, 0) + 1

    popular_books = []
    for book_id, count in sorted(book_borrow_count.items(), key=lambda x: x[1], reverse=True)[:5]:
        book = books.get(book_id)
        if book:
            popular_books.append({'book': book, 'borrow_count': count})

    # User borrowing statistics
    user_borrow_count = {}
    for transaction in transactions.values():
        user_id = transaction.user_id
        user_borrow_count[user_id] = user_borrow_count.get(user_id, 0) + 1

    active_users = []
    for user_id, count in sorted(user_borrow_count.items(), key=lambda x: x[1], reverse=True)[:5]:
        user = users.get(user_id)
        if user and user.role in ['student', 'teacher']:
            active_users.append({'user': user, 'borrow_count': count})

    # Recent activity (last 30 days)
    from datetime import timedelta
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_transactions = [t for t in transactions.values() if t.transaction_date >= thirty_days_ago]

    report_data = {
        'total_books': total_books,
        'total_copies': total_copies,
        'available_copies': available_copies,
        'borrowed_copies': borrowed_copies,
        'total_users': total_users,
        'active_borrowers': active_borrowers,
        'overdue_count': len(overdue_transactions),
        'total_fines_due': total_fines_due,
        'total_fines_collected': total_fines_collected,
        'total_revenue': total_revenue,
        'todays_collections': todays_collections,
        'overdue_transactions': overdue_transactions,
        'popular_books': popular_books,
        'active_users': active_users,
        'recent_activity_count': len(recent_transactions),
        'total_transactions': len(transactions)
    }

    return render_template('admin_reports.html', data=report_data)

@app.route('/admin/book-requests')
@login_required
def manage_book_requests():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    # Get all pending book requests with user and book details
    request_details = []
    for book_request in book_requests.values():
        user = users.get(book_request.user_id)
        book = books.get(book_request.book_id)
        request_details.append({
            'request': book_request,
            'user': user,
            'book': book
        })

    # Sort by request date (newest first)
    request_details.sort(key=lambda x: x['request'].request_date, reverse=True)

    return render_template('manage_requests.html', requests=request_details)

@app.route('/admin/book-requests/approve/<request_id>')
@login_required
def approve_book_request(request_id):
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    book_request = book_requests.get(request_id)
    if not book_request:
        flash('Request not found!', 'danger')
        return redirect(url_for('manage_book_requests'))

    if book_request.status != 'pending':
        flash('Request has already been processed!', 'warning')
        return redirect(url_for('manage_book_requests'))

    book = books.get(book_request.book_id)
    user = users.get(book_request.user_id)

    if not book or not user:
        flash('Invalid book or user!', 'danger')
        return redirect(url_for('manage_book_requests'))

    if book.available_copies <= 0:
        flash('Book is not available for borrowing!', 'warning')
        return redirect(url_for('manage_book_requests'))

    # Create transaction
    transaction_id = f"trans{len(transactions) + 1}"
    transaction = Transaction(transaction_id, book_request.user_id, book_request.book_id, 'borrow')
    transactions[transaction_id] = transaction

    # Update book availability
    book.available_copies -= 1

    # Update request status
    book_request.status = 'approved'
    book_request.processed_date = datetime.now()
    book_request.processed_by = current_user.id
    book_request.admin_response = 'Request approved and book assigned'

    log_activity(current_user.id, 'approve_book_request', get_client_ip(), {
        'book_title': book.title,
        'assigned_to': user.full_name,
        'user_role': user.role,
        'request_id': request_id
    })

    flash(f'Request approved! Book "{book.title}" assigned to {user.full_name}. Due date: {transaction.due_date.strftime("%Y-%m-%d")}', 'success')
    return redirect(url_for('manage_book_requests'))

@app.route('/admin/book-requests/reject/<request_id>', methods=['POST'])
@login_required
def reject_book_request(request_id):
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    book_request = book_requests.get(request_id)
    if not book_request:
        flash('Request not found!', 'danger')
        return redirect(url_for('manage_book_requests'))

    if book_request.status != 'pending':
        flash('Request has already been processed!', 'warning')
        return redirect(url_for('manage_book_requests'))

    rejection_reason = request.form.get('reason', 'No reason provided')

    # Update request status
    book_request.status = 'rejected'
    book_request.processed_date = datetime.now()
    book_request.processed_by = current_user.id
    book_request.admin_response = rejection_reason

    user = users.get(book_request.user_id)
    book = books.get(book_request.book_id)

    log_activity(current_user.id, 'reject_book_request', get_client_ip(), {
        'book_title': book.title if book else 'Unknown',
        'user': user.full_name if user else 'Unknown',
        'reason': rejection_reason,
        'request_id': request_id
    })

    flash(f'Request rejected successfully.', 'info')
    return redirect(url_for('manage_book_requests'))

@app.route('/admin/import-books')
@login_required
def import_books():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    return render_template('import_books.html')

@app.route('/admin/download-template')
@login_required
def download_template():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    import pandas as pd
    import os
    from io import BytesIO

    # Create sample data for the template
    template_data = {
        'Title': ['Sample Book 1', 'Sample Book 2', 'Sample Book 3'],
        'Author': ['John Doe', 'Jane Smith', 'Bob Johnson'],
        'ISBN': ['978-1234567890', '978-0987654321', '978-1122334455'],
        'Genre': ['Fiction', 'Science', 'History'],
        'Publication_Year': [2020, 2019, 2021],
        'Total_Copies': [5, 3, 2]
    }

    df = pd.DataFrame(template_data)

    # Create a BytesIO object to store the Excel file in memory
    output = BytesIO()

    # Create an Excel writer object with styling
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Books', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Books']

        # Add some styling
        from openpyxl.styles import Font, PatternFill

        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='book_import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/import-books', methods=['POST'])
@login_required
def import_books_post():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    if 'excel_file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('import_books'))

    file = request.files['excel_file']

    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('import_books'))

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)!', 'danger')
        return redirect(url_for('import_books'))

    try:
        import pandas as pd

        # Read the Excel file
        df = pd.read_excel(file)

        # Validate required columns
        required_columns = ['Title', 'Author', 'ISBN', 'Genre', 'Publication_Year', 'Total_Copies']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
            return redirect(url_for('import_books'))

        # Process each row
        imported_count = 0
        error_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # Validate data
                title = str(row['Title']).strip()
                author = str(row['Author']).strip()
                isbn = str(row['ISBN']).strip()
                genre = str(row['Genre']).strip()

                if not title or not author or not isbn:
                    errors.append(f"Row {index + 2}: Missing required data (Title, Author, or ISBN)")
                    error_count += 1
                    continue

                # Check if book already exists (by ISBN)
                existing_book = None
                for book in books.values():
                    if book.isbn == isbn:
                        existing_book = book
                        break

                if existing_book:
                    errors.append(f"Row {index + 2}: Book with ISBN {isbn} already exists")
                    error_count += 1
                    continue

                # Convert and validate numeric fields
                try:
                    publication_year = int(row['Publication_Year'])
                    total_copies = int(row['Total_Copies'])

                    if publication_year < 1000 or publication_year > 2030:
                        errors.append(f"Row {index + 2}: Invalid publication year: {publication_year}")
                        error_count += 1
                        continue

                    if total_copies < 1:
                        errors.append(f"Row {index + 2}: Total copies must be at least 1")
                        error_count += 1
                        continue

                except ValueError:
                    errors.append(f"Row {index + 2}: Invalid numeric data (Publication Year or Total Copies)")
                    error_count += 1
                    continue

                # Create new book
                book_id = f"book{len(books) + 1}"
                book = Book(book_id, title, author, isbn, genre, publication_year, total_copies)
                books[book_id] = book
                imported_count += 1

            except Exception as e:
                errors.append(f"Row {index + 2}: Error processing data - {str(e)}")
                error_count += 1

        # Log the import activity
        log_activity(current_user.id, 'import_books', get_client_ip(), {
            'imported_count': imported_count,
            'error_count': error_count,
            'total_rows': len(df)
        })

        # Prepare flash messages
        if imported_count > 0:
            flash(f'Successfully imported {imported_count} books!', 'success')

        if error_count > 0:
            flash(f'{error_count} rows had errors. See details below.', 'warning')
            # Store errors in session for display
            session['import_errors'] = errors[:10]  # Limit to first 10 errors

        return redirect(url_for('import_books'))

    except Exception as e:
        flash(f'Error processing Excel file: {str(e)}', 'danger')
        return redirect(url_for('import_books'))

@app.route('/admin/import-users')
@login_required
def import_users():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    return render_template('import_users.html')

@app.route('/admin/download-user-template')
@login_required
def download_user_template():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    import pandas as pd
    import os
    from io import BytesIO

    # Create sample data for the template
    template_data = {
        'Full_Name': ['John Doe', 'Jane Smith', 'Bob Johnson'],
        'Email': ['john.doe@example.com', 'jane.smith@example.com', 'bob.johnson@example.com'],
        'Role': ['student', 'teacher', 'student'],
        'Admission_Number': ['2024001', '', '2024002'],  # Only for students
        'Password': ['student123', 'teacher123', 'student456']
    }

    df = pd.DataFrame(template_data)

    # Create a BytesIO object to store the Excel file in memory
    output = BytesIO()

    # Create an Excel writer object with styling
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Users', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Users']

        # Add some styling
        from openpyxl.styles import Font, PatternFill

        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='user_import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/import-users', methods=['POST'])
@login_required
def import_users_post():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    if 'excel_file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('import_users'))

    file = request.files['excel_file']

    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('import_users'))

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('Please upload an Excel file (.xlsx or .xls)!', 'danger')
        return redirect(url_for('import_users'))

    try:
        import pandas as pd

        # Read the Excel file
        df = pd.read_excel(file)

        # Validate required columns
        required_columns = ['Full_Name', 'Email', 'Role', 'Password']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
            return redirect(url_for('import_users'))

        # Process each row
        imported_count = 0
        error_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # Validate data
                full_name = str(row['Full_Name']).strip()
                email = str(row['Email']).strip()
                role = str(row['Role']).strip().lower()
                password = str(row['Password']).strip()
                admission_number = str(row.get('Admission_Number', '')).strip() if 'Admission_Number' in row else ''

                if not full_name or not email or not role or not password:
                    errors.append(f"Row {index + 2}: Missing required data")
                    error_count += 1
                    continue

                if role not in ['student', 'teacher']:
                    errors.append(f"Row {index + 2}: Invalid role '{role}'. Must be 'student' or 'teacher'")
                    error_count += 1
                    continue

                # For students, admission number is required and used as username
                if role == 'student':
                    if not admission_number:
                        errors.append(f"Row {index + 2}: Admission number required for students")
                        error_count += 1
                        continue
                    username = admission_number
                else:
                    # For teachers, generate username from email
                    username = email.split('@')[0]

                # Check if username already exists
                if User.get_by_username(username):
                    errors.append(f"Row {index + 2}: Username '{username}' already exists")
                    error_count += 1
                    continue

                # Create new user
                user_id = f"{role}{len([u for u in users.values() if u.role == role]) + 1}"
                user = User(user_id, username, email, password, role, full_name, admission_number if role == 'student' else None)
                users[user_id] = user
                imported_count += 1

            except Exception as e:
                errors.append(f"Row {index + 2}: Error processing data - {str(e)}")
                error_count += 1

        # Log the import activity
        log_activity(current_user.id, 'import_users', get_client_ip(), {
            'imported_count': imported_count,
            'error_count': error_count,
            'total_rows': len(df)
        })

        # Prepare flash messages
        if imported_count > 0:
            flash(f'Successfully imported {imported_count} users!', 'success')

        if error_count > 0:
            flash(f'{error_count} rows had errors. See details below.', 'warning')
            # Store errors in session for display
            session['import_errors'] = errors[:10]  # Limit to first 10 errors

        return redirect(url_for('import_users'))

    except Exception as e:
        flash(f'Error processing Excel file: {str(e)}', 'danger')
        return redirect(url_for('import_users'))

@app.route('/admin/export-financial-report')
@login_required
def export_financial_report():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    import pandas as pd
    from io import BytesIO

    # Prepare financial data
    financial_data = []

    # All transactions with fines
    for transaction in transactions.values():
        transaction.update_fine()
        if transaction.fine_amount > 0:
            user = users.get(transaction.user_id)
            book = books.get(transaction.book_id)

            # Find corresponding payment record
            payment_record = None
            for payment in payment_records.values():
                if payment.transaction_id == transaction.id:
                    payment_record = payment
                    break

            financial_data.append({
                'Transaction_ID': transaction.id,
                'User_Name': user.full_name if user else 'Unknown',
                'User_Role': user.role.title() if user else 'Unknown',
                'Admission_Number': user.admission_number if user and user.admission_number else 'N/A',
                'Book_Title': book.title if book else 'Unknown',
                'Book_Author': book.author if book else 'Unknown',
                'Borrow_Date': transaction.transaction_date.strftime('%Y-%m-%d'),
                'Due_Date': transaction.due_date.strftime('%Y-%m-%d'),
                'Return_Date': transaction.return_date.strftime('%Y-%m-%d') if transaction.return_date else 'Not Returned',
                'Days_Overdue': transaction.days_overdue,
                'Fine_Amount': transaction.fine_amount,
                'Fine_Status': 'Paid' if transaction.fine_paid else 'Pending',
                'Payment_Date': payment_record.payment_date.strftime('%Y-%m-%d %H:%M') if payment_record else 'N/A',
                'Payment_Method': payment_record.payment_method if payment_record else 'N/A',
                'Collected_By': users.get(payment_record.collected_by).full_name if payment_record and payment_record.collected_by else 'N/A',
                'Payment_Notes': payment_record.notes if payment_record else 'N/A'
            })

    if not financial_data:
        flash('No financial data available to export!', 'warning')
        return redirect(url_for('admin_reports'))

    # Create DataFrame
    df = pd.DataFrame(financial_data)

    # Create summary statistics
    total_fines = df['Fine_Amount'].sum()
    paid_fines = df[df['Fine_Status'] == 'Paid']['Fine_Amount'].sum()
    pending_fines = df[df['Fine_Status'] == 'Pending']['Fine_Amount'].sum()

    summary_data = {
        'Metric': ['Total Fines Generated', 'Fines Collected', 'Fines Pending', 'Collection Rate (%)'],
        'Amount (rs)': [total_fines, paid_fines, pending_fines, f"{(paid_fines/total_fines*100):.1f}%" if total_fines > 0 else "0%"]
    }
    summary_df = pd.DataFrame(summary_data)

    # Create Excel file with multiple sheets
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Financial details sheet
        df.to_excel(writer, sheet_name='Financial_Details', index=False)

        # Summary sheet
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Style the workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Style financial details sheet
        workbook = writer.book
        worksheet1 = writer.sheets['Financial_Details']
        worksheet2 = writer.sheets['Summary']

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

        for sheet in [worksheet1, worksheet2]:
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

        # Auto-adjust column widths
        for sheet in [worksheet1, worksheet2]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Color code fine status in financial details
        for row in range(2, len(financial_data) + 2):
            status_cell = worksheet1[f'L{row}']  # Fine_Status column
            if status_cell.value == 'Paid':
                status_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    output.seek(0)

    # Log export activity
    log_activity(current_user.id, 'export_financial_report', get_client_ip(), {
        'total_records': len(financial_data),
        'total_amount': f"{total_fines} rs",
        'export_format': 'xlsx'
    })

    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    return send_file(
        output,
        as_attachment=True,
        download_name=f'library_financial_report_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/export-fine-collections-pdf')
@login_required
def export_fine_collections_pdf():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    from datetime import datetime, date
    from flask import make_response
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    import calendar

    # Get date range from query parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = date.today().replace(day=1)  # First day of current month

        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = date.today()
    except ValueError:
        flash('Invalid date format! Please use YYYY-MM-DD format.', 'danger')
        return redirect(url_for('admin_reports'))

    # Filter payment records by date range
    filtered_payments = []
    for payment in payment_records.values():
        payment_date = payment.payment_date.date()
        if start_date <= payment_date <= end_date:
            transaction = transactions.get(payment.transaction_id)
            user = users.get(payment.user_id)
            book = None
            if transaction:
                book = books.get(transaction.book_id)

            filtered_payments.append({
                'payment': payment,
                'transaction': transaction,
                'user': user,
                'book': book
            })

    if not filtered_payments:
        flash(f'No fine collections found between {start_date} and {end_date}!', 'warning')
        return redirect(url_for('admin_reports'))

    # Sort by payment date
    filtered_payments.sort(key=lambda x: x['payment'].payment_date)

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

    # Container for the 'Flowable' objects
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )

    # Add title
    title = Paragraph("Little Flower Library - Fine Collections Report", title_style)
    elements.append(title)

    # Add date range
    date_range = Paragraph(f"<b>Period:</b> {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}", styles['Normal'])
    elements.append(date_range)
    elements.append(Spacer(1, 20))

    # Calculate summary
    total_amount = sum(item['payment'].amount for item in filtered_payments)
    total_transactions = len(filtered_payments)

    # Add summary
    summary_data = [
        ['Summary', ''],
        ['Total Collections:', f'{total_amount:.2f} rs'],
        ['Total Transactions:', str(total_transactions)],
        ['Average per Transaction:', f'{total_amount/total_transactions:.2f} rs' if total_transactions > 0 else '0.00 rs']
    ]

    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 30))

    # Add detailed collections table
    details_title = Paragraph("<b>Detailed Collections</b>", styles['Heading2'])
    elements.append(details_title)
    elements.append(Spacer(1, 10))

    # Prepare data for the table
    data = [['Date', 'Student/Teacher', 'Book Title', 'Amount (rs)', 'Payment Method']]

    for item in filtered_payments:
        payment = item['payment']
        user = item['user']
        book = item['book']

        data.append([
            payment.payment_date.strftime('%Y-%m-%d'),
            user.full_name if user else 'Unknown',
            book.title if book else 'Unknown Book',
            f'{payment.amount:.2f}',
            payment.payment_method.title()
        ])

    # Create table
    table = Table(data, colWidths=[1.2*inch, 1.8*inch, 2*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
    ]))

    elements.append(table)

    # Add footer
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1  # Center alignment
    )
    footer = Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", footer_style)
    elements.append(footer)

    # Build PDF
    doc.build(elements)

    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()

    # Log export activity
    log_activity(current_user.id, 'export_fine_collections_pdf', get_client_ip(), {
        'date_range': f'{start_date} to {end_date}',
        'total_records': len(filtered_payments),
        'total_amount': f'{total_amount:.2f} rs'
    })

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=fine_collections_{start_date}_{end_date}.pdf'

    return response

@app.route('/admin/fine-collections-report')
@login_required
def fine_collections_report():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    from datetime import datetime, timedelta
    return render_template('fine_collections_report.html', datetime=datetime, timedelta=timedelta)

@app.route('/admin/export-books')
@login_required
def export_books():
    if current_user.role != 'admin':
        flash('Access denied! Admin privileges required.', 'danger')
        return redirect(url_for('index'))

    import pandas as pd
    from io import BytesIO

    # Prepare book data for export
    book_data = []
    for book in books.values():
        book_data.append({
            'Title': book.title,
            'Author': book.author,
            'ISBN': book.isbn,
            'Genre': book.genre,
            'Publication_Year': book.publication_year,
            'Total_Copies': book.total_copies,
            'Available_Copies': book.available_copies,
            'Borrowed_Copies': book.total_copies - book.available_copies
        })

    if not book_data:
        flash('No books available to export!', 'warning')
        return redirect(url_for('book_management'))

    df = pd.DataFrame(book_data)

    # Create Excel file with styling
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Library_Books', index=False)

        # Get workbook and worksheet for styling
        workbook = writer.book
        worksheet = writer.sheets['Library_Books']

        # Add styling
        from openpyxl.styles import Font, PatternFill

        # Style header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    output.seek(0)

    # Log export activity
    log_activity(current_user.id, 'export_books', get_client_ip(), {
        'exported_count': len(book_data),
        'export_format': 'xlsx'
    })

    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    return send_file(
        output,
        as_attachment=True,
        download_name=f'library_books_export_{timestamp}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/developer')
def developer_login():
    return render_template('developer_login.html')

@app.route('/developer/login', methods=['POST'])
def developer_login_post():
    username = request.form['username']
    password = request.form['password']
    
    # Developer credentials (in production, use environment variables)
    if username == 'developer' and password == 'dev123':
        session['developer_authenticated'] = True
        log_activity('developer', 'developer_login', get_client_ip())
        flash('Developer login successful!', 'success')
        return redirect(url_for('developer_dashboard'))
    else:
        flash('Invalid developer credentials!', 'danger')
        return redirect(url_for('developer_login'))

@app.route('/developer/dashboard')
def developer_dashboard():
    if not session.get('developer_authenticated'):
        flash('Developer authentication required!', 'danger')
        return redirect(url_for('developer_login'))
    
    from models import system_settings
    return render_template('developer_dashboard.html', settings=system_settings)

@app.route('/developer/settings', methods=['POST'])
def update_system_settings():
    if not session.get('developer_authenticated'):
        flash('Developer authentication required!', 'danger')
        return redirect(url_for('developer_login'))
    
    from models import system_settings
    
    # Update settings
    system_settings.school_name = request.form['school_name']
    system_settings.library_name = request.form['library_name']
    system_settings.primary_color = request.form['primary_color']
    system_settings.secondary_color = request.form['secondary_color']
    system_settings.accent_color = request.form['accent_color']
    system_settings.logo_url = request.form['logo_url']
    system_settings.developer_name = request.form['developer_name']
    system_settings.footer_text = request.form['footer_text']
    system_settings.address = request.form['address']
    system_settings.phone = request.form['phone']
    system_settings.email = request.form['email']
    
    log_activity('developer', 'update_system_settings', get_client_ip(), {
        'school_name': system_settings.school_name,
        'library_name': system_settings.library_name
    })
    
    flash('System settings updated successfully!', 'success')
    return redirect(url_for('developer_dashboard'))

@app.route('/developer/logout')
def developer_logout():
    session.pop('developer_authenticated', None)
    flash('Developer logged out successfully!', 'info')
    return redirect(url_for('index'))